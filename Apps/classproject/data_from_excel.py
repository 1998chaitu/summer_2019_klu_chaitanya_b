import os
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'classproject.settings')
import django
django.setup()
from onlineapp.models import *
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import click

@click.command()
@click.argument('args',nargs=2)
def cli(args):
    wb = load_workbook(filename=args[0])
    #print(wb.sheetnames)
    data=wb[args[1]]
    print(args[0],args[1])
    return
    flag=0
    for row in data.rows:
        if flag==0:
            flag=1
            continue
        vals=[x.value for x in row]
        if args[1]=='Colleges':
            c=College(name=vals[0],location=vals[2],acronym=vals[1],contact=vals[3])
        elif args[1]=='Current':
            c = Student(name=vals[0], email=vals[2], db_folder=vals[3],college=College.objects.get(acronym=vals[1]),dropped_out=False)
        elif args[1]=='Deletions':
            c = Student(name=vals[0], email=vals[2], db_folder=vals[3], college=College.objects.get(acronym=vals[1]),dropped_out=True)
        elif args[1]=='Marks':
            stu=vals[0].split('_')[2]
            p1,p2,p3,p4=int(vals[1]),int(vals[2]),int(vals[3]),int(vals[4])
            try:
                folder=Student.objects.get(db_folder=stu)
            except:
                folder=None
            c=MockTest1(problem1=p1,problem2=p2,problem3=p3,problem4=p4,total=vals[5],student=folder)

        c.save()



if __name__ == '__main__':
    cli()