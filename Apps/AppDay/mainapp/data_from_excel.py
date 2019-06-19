import os
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'mainapp.settings')
import django
django.setup()
from app.models import *
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import click
import datetime


@click.command()
@click.argument('args',nargs=2)
def cli(args):
    wb = load_workbook(filename=args[0])
    #print(wb.sheetnames)
    print(args[0],args[1])
    #return
    data = wb['deliveries']
    flag=0
    for row in data.rows:
        if flag==0:
            flag=1
            continue
        vals=[x.value for x in row]
        # if args[1] == 'Matches':
        #     print('123')
        #     m = Matches(id=vals[0], season=vals[1], city=vals[2], date=vals[3],  team1=vals[4], team2=vals[5], toss_winner=vals[6], toss_decision=vals[7], result=vals[8], dl_applied=vals[9], winner=vals[10], win_by_runs=vals[11], win_by_wickets=vals[12], player_of_match=vals[13], venue=vals[14], umpire1=vals[15], umpire2=vals[16], umpire3=vals[17])
        #     m.save()
        # for i in range(18):
        #     print(vals[i],end=' ')
        # print()

        if args[1] == 'D':
            print('123')
            d = Deliveries(match_id=Matches.objects.get(id=vals[0]), inning=vals[1], batting_team=vals[2], bowling_team=vals[3], over=vals[4], ball=vals[5], batsman=vals[6], non_stricker=vals[7], bowler=vals[8], is_super_over=vals[9], wide_runs=vals[10], bye_runs=vals[11], legbye_runs=vals[12], noball_runs=vals[13], penality_runs=vals[14], batsman_runs=vals[15], extra_runs=vals[16], total_runs=vals[17], player_dismissed=vals[18], dismissal_kind=vals[19], fielder=vals[20])
            d.save()

if __name__ == '__main__':
    cli()